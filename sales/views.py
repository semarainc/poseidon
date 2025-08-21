from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, Response
from datetime import datetime, date, timedelta
from sqlalchemy import func, or_, inspect as sqlalchemy_inspect
import os
import uuid
import io
import csv
from werkzeug.utils import secure_filename
import traceback
from flask_login import login_required, current_user
from models.models import db, Barang, Pelanggan, Penjualan, DetailPenjualan, generate_no_transaksi, PenjualanAntrian
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
sales_blueprint = Blueprint('sales', __name__, template_folder='templates', static_folder='static', url_prefix='/sales')


# --- PENJUALAN ---

# === ANTRIAN PENJUALAN (PER USER) ===
from flask import jsonify

@sales_blueprint.route('/antrian', methods=['GET'], strict_slashes=False)
@login_required
def list_antrian():
    antrians = PenjualanAntrian.query.filter_by(user_id=current_user.id).order_by(PenjualanAntrian.created_at).all()
    return jsonify([
        {
            'id': a.id,
            'nama_antrian': a.nama_antrian,
            'created_at': a.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': a.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
        } for a in antrians
    ])

@sales_blueprint.route('/antrian/tambah', methods=['POST'], strict_slashes=False)
@login_required
def tambah_antrian():
    data = request.get_json()
    nama_antrian = data.get('nama_antrian', f'Antrian {datetime.now().strftime("%H%M%S")}')
    data_antrian = data.get('data_antrian')
    # Perbolehkan data_antrian berupa string "[]" (array kosong) atau string kosong
    if data_antrian is None:
        return jsonify({'success': False, 'message': 'Data antrian kosong'}), 400
    antrian = PenjualanAntrian(user_id=current_user.id, nama_antrian=nama_antrian, data_antrian=data_antrian)
    db.session.add(antrian)
    db.session.commit()
    return jsonify({'success': True, 'id': antrian.id, 'nama_antrian': antrian.nama_antrian})

@sales_blueprint.route('/antrian/<int:antrian_id>/ambil', methods=['GET'], strict_slashes=False)
@login_required
def ambil_antrian(antrian_id):
    antrian = PenjualanAntrian.query.filter_by(id=antrian_id, user_id=current_user.id).first_or_404()
    return jsonify({'id': antrian.id, 'nama_antrian': antrian.nama_antrian, 'data_antrian': antrian.data_antrian})

@sales_blueprint.route('/antrian/<int:antrian_id>/hapus', methods=['POST'])
@login_required
def hapus_antrian(antrian_id):
    antrian = PenjualanAntrian.query.filter_by(id=antrian_id, user_id=current_user.id).first_or_404()
    db.session.delete(antrian)
    db.session.commit()
    return jsonify({'success': True})

@sales_blueprint.route('/antrian/<int:antrian_id>/update', methods=['POST'], strict_slashes=False)
@login_required
def update_antrian(antrian_id):
    """Update nama_antrian atau data_antrian milik pengguna saat ini."""
    antrian = PenjualanAntrian.query.filter_by(id=antrian_id, user_id=current_user.id).first_or_404()
    data = request.get_json() or {}
    if 'nama_antrian' in data and data['nama_antrian']:
        antrian.nama_antrian = data['nama_antrian']
    if 'data_antrian' in data:
        antrian.data_antrian = data['data_antrian']
    antrian.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'success': True})

@sales_blueprint.route('/penjualan')
@login_required
def penjualan():
    pelanggan_list = Pelanggan.query.order_by(Pelanggan.nama).all()
    
    # Read view mode preference from cookie
    view_mode = request.cookies.get('sales_view_mode', 'modern')
    
    # Render appropriate template based on preference
    if view_mode == 'classic':
        return render_template('penjualan_classic.html', pelanggan_list=pelanggan_list)
    else:
        return render_template('penjualan.html', pelanggan_list=pelanggan_list)

@sales_blueprint.route('/api/cari_barang', strict_slashes=False)
@login_required
def cari_barang_api():
    query_param = request.args.get('q', '')
    limit = request.args.get('limit', 10, type=int)
    sort = request.args.get('sort', 'name')  # 'name' or 'popular'

    # Always try to prioritize exact kode match if query is given
    exact_item = None
    if query_param:
        # Case-insensitive exact match on kode_barang
        exact_item = Barang.query.filter(Barang.kode_barang.ilike(query_param)).first()

    remaining_limit = limit
    ordered_items = []

    if exact_item:
        ordered_items.append(exact_item)
        remaining_limit = max(0, limit - 1)

    if remaining_limit > 0:
        if sort == 'popular':
            # Popularity based on total quantity sold in DetailPenjualan
            qty_sum = func.coalesce(func.sum(DetailPenjualan.jumlah), 0).label('total_qty')
            query_builder = db.session.query(Barang, qty_sum).outerjoin(DetailPenjualan, DetailPenjualan.barang_id == Barang.id)
            if query_param:
                query_param_like = f"%{query_param}%"
                query_builder = query_builder.filter(
                    or_(
                        Barang.nama_barang.ilike(query_param_like),
                        Barang.kode_barang.ilike(query_param_like)
                    )
                )
            if exact_item:
                query_builder = query_builder.filter(Barang.id != exact_item.id)
            query_builder = query_builder.group_by(Barang.id).order_by(qty_sum.desc())
            rows = query_builder.limit(remaining_limit).all()
            ordered_items.extend([row[0] for row in rows])
        else:
            query_builder = Barang.query
            if query_param:
                query_param_like = f"%{query_param}%"
                query_builder = query_builder.filter(
                    or_(
                        Barang.nama_barang.ilike(query_param_like),
                        Barang.kode_barang.ilike(query_param_like)
                    )
                )
            if exact_item:
                query_builder = query_builder.filter(Barang.id != exact_item.id)
            ordered_items.extend(query_builder.order_by(Barang.nama_barang).limit(remaining_limit).all())

    results = [item.to_dict() for item in ordered_items]
    return jsonify(results)

@sales_blueprint.route('/proses_penjualan', methods=['POST'], strict_slashes=False)
@login_required
def proses_penjualan():
    data = request.get_json()
    if not data or 'items' not in data or not data['items']:
        return jsonify({'success': False, 'message': 'Tidak ada item dalam transaksi.'}), 400

    # ---Cegah duplikasi transaksi ---
    queue_id = data.get('queue_id')  # Pastikan frontend mengirim queue_id (backendId antrean)
    if queue_id:
        # Cek apakah sudah ada transaksi untuk queue/antrian ini
        existing_penjualan = PenjualanAntrian.query.filter_by(id=queue_id, user_id=current_user.id).first()
        if not existing_penjualan:
            # Sudah dihapus, tidak boleh diproses ulang
            return jsonify({'success': False, 'message': 'Transaksi ini sudah diproses atau antrean sudah dihapus.'}), 409
    
    no_transaksi = generate_no_transaksi()
    pelanggan_id_str = data.get('pelanggan_id')
    metode_pembayaran_frontend = data.get('metode_pembayaran', 'tunai')
    jumlah_bayar_tunai_frontend = data.get('jumlah_bayar_tunai') 

    pelanggan_id = None
    if pelanggan_id_str and pelanggan_id_str.isdigit():
        pelanggan_id = int(pelanggan_id_str)
        if pelanggan_id is not None and not Pelanggan.query.get(pelanggan_id):
             current_app.logger.warning(f"Pelanggan ID {pelanggan_id} provided but not found. Defaulting to null.")
             pelanggan_id = None

    try:
        total_harga_frontend = float(data['total_harga'])
        
        calculated_total_harga = 0
        for item_data_val in data['items']:
            if not all(k in item_data_val for k in ['barang_id', 'jumlah', 'harga_satuan']):
                db.session.rollback()
                return jsonify({'success': False, 'message': 'Data item tidak lengkap.'}), 400
            
            barang_to_calc = Barang.query.get(item_data_val['barang_id'])
            if not barang_to_calc:
                 db.session.rollback()
                 return jsonify({'success': False, 'message': f"Item barang dengan ID {item_data_val['barang_id']} tidak ditemukan untuk kalkulasi."}), 404
            
            harga_satuan_val = float(item_data_val['harga_satuan']) 
            jumlah_val = int(item_data_val['jumlah'])
            calculated_total_harga += harga_satuan_val * jumlah_val
        
        final_total_harga = round(calculated_total_harga, 2)

        if abs(final_total_harga - total_harga_frontend) > 0.01: 
            current_app.logger.warning(f"Total harga mismatch: Frontend {total_harga_frontend}, Backend {final_total_harga}. Using backend calculated total.")

        db_jumlah_bayar = None
        db_kembalian = None

        if metode_pembayaran_frontend == 'tunai':
            if jumlah_bayar_tunai_frontend is None:
                return jsonify({'success': False, 'message': 'Jumlah bayar tunai diperlukan untuk metode tunai.'}), 400
            try:
                db_jumlah_bayar = float(jumlah_bayar_tunai_frontend)
            except ValueError:
                 return jsonify({'success': False, 'message': 'Format jumlah bayar tunai tidak valid.'}), 400
            
            if db_jumlah_bayar < final_total_harga:
                return jsonify({'success': False, 'message': f'Jumlah bayar (Rp {db_jumlah_bayar:,.2f}) kurang dari total harga (Rp {final_total_harga:,.2f}).'}), 400
            db_kembalian = round(db_jumlah_bayar - final_total_harga, 2)
        else: 
            db_jumlah_bayar = final_total_harga
            db_kembalian = 0.0


        new_penjualan = Penjualan(
            no_transaksi=no_transaksi,
            pelanggan_id=pelanggan_id,
            total_harga=final_total_harga,
            metode_pembayaran=metode_pembayaran_frontend,
            jumlah_bayar=db_jumlah_bayar,
            kembalian=db_kembalian
        )
        db.session.add(new_penjualan)
        db.session.flush() 
        
        items_to_process = data['items']
        processed_details = []

        for item_data in items_to_process:
            barang_id_tobuy = item_data['barang_id']
            jumlah_beli = int(item_data['jumlah'])
            barang_db_check = Barang.query.get(barang_id_tobuy)
            if not barang_db_check:
                db.session.rollback()
                return jsonify({'success': False, 'message': f"Barang dengan ID {barang_id_tobuy} tidak ditemukan."}), 404

            harga_saat_transaksi = barang_db_check.harga_jual 
            tipe_harga = 'eceran'
            
            # Cek apakah memenuhi syarat harga grosir
            if barang_db_check.harga_grosir is not None and jumlah_beli >= (barang_db_check.batas_minimal_grosir or 10):
                harga_saat_transaksi = barang_db_check.harga_grosir
                tipe_harga = 'grosir'
                
            subtotal_item = round(harga_saat_transaksi * jumlah_beli, 2)

            updated_rows = db.session.query(Barang).filter(
                Barang.id == barang_id_tobuy,
                Barang.stok >= jumlah_beli  
            ).update({"stok": Barang.stok - jumlah_beli}, synchronize_session='fetch') 

            if updated_rows == 0: 
                db.session.rollback() 
                barang_aktual_stok = Barang.query.get(barang_id_tobuy) 
                message = ""
                if barang_aktual_stok: 
                    stok_aktual = barang_aktual_stok.stok
                    nama_barang_aktual = barang_aktual_stok.nama_barang
                    message = f'Stok {nama_barang_aktual} tidak mencukupi (tersisa: {stok_aktual}). Transaksi dibatalkan.'
                else: 
                    message = f"Barang dengan ID {barang_id_tobuy} tidak ditemukan atau stok tidak dapat diverifikasi. Transaksi dibatalkan."
                return jsonify({'success': False, 'message': message }), 409 
            
            detail = DetailPenjualan(
                penjualan_id=new_penjualan.id,
                barang_id=barang_id_tobuy,
                jumlah=jumlah_beli,
                harga_satuan=harga_saat_transaksi, 
                subtotal=subtotal_item,
                tipe_harga=tipe_harga
            )
            processed_details.append(detail)
        
        if processed_details:
            db.session.add_all(processed_details)
        
        db.session.commit() 
        return jsonify({
            'success': True, 
            'no_transaksi': no_transaksi, 
            'penjualan_id': new_penjualan.id,
            'kembalian_tunai': db_kembalian if metode_pembayaran_frontend == 'tunai' else None 
            })

    except ValueError as ve: 
        db.session.rollback()
        traceback.print_exc()
        current_app.logger.error(f"ValueError during sale processing: {ve}", exc_info=True)
        return jsonify({'success': False, 'message': f'Data transaksi tidak valid: {str(ve)}'}), 400
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        current_app.logger.error(f"Error processing sale: {e}", exc_info=True) 
        return jsonify({'success': False, 'message': f'Terjadi kesalahan server saat memproses penjualan. Silakan coba lagi.'}), 500


@sales_blueprint.route('/struk/<int:penjualan_id>', strict_slashes=False)
@login_required
def struk(penjualan_id):
    penjualan_obj = Penjualan.query.options(
        db.joinedload(Penjualan.pelanggan_data), 
        db.joinedload(Penjualan.detail_penjualan).joinedload(DetailPenjualan.barang_data) 
    ).get_or_404(penjualan_id)
    
    return render_template('struk.html', 
                           penjualan=penjualan_obj,
                           pelanggan=penjualan_obj.pelanggan_data, 
                           detail_penjualan=penjualan_obj.detail_penjualan)

# --- RIWAYAT PENJUALAN ---
@sales_blueprint.route('/riwayat_penjualan', strict_slashes=False)
@login_required
def riwayat_penjualan():
    page = request.args.get('page', 1, type=int)
    per_page = 10 
    
    penjualan_list = Penjualan.query.options(db.joinedload(Penjualan.pelanggan_data)) \
                                    .order_by(Penjualan.created_at.desc()) \
                                    .paginate(page=page, per_page=per_page, error_out=False)
    
    return render_template('riwayat_penjualan.html', penjualan_list=penjualan_list)


# --- LAPORAN PENJUALAN ---
@sales_blueprint.route('/laporan_penjualan', methods=['GET'], strict_slashes=False)
@login_required
def laporan_penjualan():
    tanggal_mulai_str = request.args.get('tanggal_mulai')
    tanggal_akhir_str = request.args.get('tanggal_akhir')
    
    query = Penjualan.query.options(
        db.joinedload(Penjualan.pelanggan_data),
        db.joinedload(Penjualan.detail_penjualan).joinedload(DetailPenjualan.barang_data) 
    ).order_by(Penjualan.tanggal.desc())

    tanggal_mulai = None
    tanggal_akhir = None

    if tanggal_mulai_str:
        try:
            tanggal_mulai = datetime.strptime(tanggal_mulai_str, '%Y-%m-%d').date()
            query = query.filter(Penjualan.tanggal >= datetime.combine(tanggal_mulai, datetime.min.time()))
        except ValueError:
            flash('Format tanggal mulai tidak valid.', 'error')
            tanggal_mulai = None 
    
    if tanggal_akhir_str:
        try:
            tanggal_akhir = datetime.strptime(tanggal_akhir_str, '%Y-%m-%d').date()
            query = query.filter(Penjualan.tanggal <= datetime.combine(tanggal_akhir, datetime.max.time()))
        except ValueError:
            flash('Format tanggal akhir tidak valid.', 'error')
            tanggal_akhir = None 

    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    allowed_sizes = [10, 25, 50, 100]
    if per_page not in allowed_sizes:
        per_page = 25
    
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    transaksi_list = pagination.items
    total_count = pagination.total
    total_penjualan_periode = sum(t.total_harga or 0.0 for t in transaksi_list)
    
    penjualan_per_hari = {}
    if isinstance(tanggal_mulai, date) and isinstance(tanggal_akhir, date):
        daily_sales_query = db.session.query(
            func.date(Penjualan.tanggal).label('tgl_penjualan'),
            func.sum(Penjualan.total_harga).label('total_harian'),
            func.count(Penjualan.id).label('jumlah_transaksi_harian')
        ).filter(Penjualan.tanggal >= datetime.combine(tanggal_mulai, datetime.min.time()), 
                 Penjualan.tanggal <= datetime.combine(tanggal_akhir, datetime.max.time())) \
         .group_by(func.date(Penjualan.tanggal)) \
         .order_by(func.date(Penjualan.tanggal)) \
         .all()
        
        for row in daily_sales_query:
            if row.tgl_penjualan: 
                penjualan_per_hari[row.tgl_penjualan] = { 
                    'total': row.total_harian or 0.0,
                    'jumlah_transaksi': row.jumlah_transaksi_harian or 0
                }
    
    return render_template('laporan_penjualan.html', 
                           transaksi_list=transaksi_list,
                           pagination=pagination,
                           total_penjualan_periode=total_penjualan_periode,
                           penjualan_per_hari=penjualan_per_hari,
                           filter_tanggal_mulai=tanggal_mulai_str if tanggal_mulai else '', 
                           filter_tanggal_akhir=tanggal_akhir_str if tanggal_akhir else '',
                           per_page=per_page,
                           page=page,
                           total_count=total_count)

@sales_blueprint.route('/laporan_penjualan/download_csv', methods=['GET'], strict_slashes=False)
@login_required
def download_laporan_csv():
    tanggal_mulai_str = request.args.get('tanggal_mulai')
    tanggal_akhir_str = request.args.get('tanggal_akhir')
    detail_item = request.args.get('detail_item', 'false').lower() == 'true'

    query = Penjualan.query.options(
        db.joinedload(Penjualan.pelanggan_data),
        db.joinedload(Penjualan.detail_penjualan).joinedload(DetailPenjualan.barang_data)
    ).order_by(Penjualan.tanggal.asc())

    if tanggal_mulai_str:
        try:
            tanggal_mulai = datetime.strptime(tanggal_mulai_str, '%Y-%m-%d').date()
            query = query.filter(Penjualan.tanggal >= datetime.combine(tanggal_mulai, datetime.min.time()))
        except ValueError: pass 
    
    if tanggal_akhir_str:
        try:
            tanggal_akhir = datetime.strptime(tanggal_akhir_str, '%Y-%m-%d').date()
            query = query.filter(Penjualan.tanggal <= datetime.combine(tanggal_akhir, datetime.max.time()))
        except ValueError: pass

    transaksi_list = query.all()
    si = io.StringIO()
    cw = csv.writer(si)

    header = ['No Transaksi', 'Tanggal', 'Waktu', 'Pelanggan', 'Total Harga', 'Metode Pembayaran', 'Jumlah Bayar', 'Kembalian', 'Status']
    if detail_item:
        header.extend(['Kode Barang', 'Nama Barang', 'Jumlah', 'Harga Satuan Item', 'Subtotal Item'])
    cw.writerow(header)

    for trx in transaksi_list:
        base_row = [
            trx.no_transaksi,
            trx.tanggal.strftime('%Y-%m-%d'),
            trx.tanggal.strftime('%H:%M:%S'),
            trx.pelanggan_data.nama if trx.pelanggan_data else 'Umum',
            trx.total_harga or 0.0, 
            trx.metode_pembayaran or '',
            trx.jumlah_bayar if trx.jumlah_bayar is not None else '',
            trx.kembalian if trx.kembalian is not None else '',
            trx.status
        ]
        if detail_item:
            if trx.detail_penjualan:
                for item_detail in trx.detail_penjualan:
                    item_row_data = [
                        item_detail.barang_data.kode_barang if item_detail.barang_data else '',
                        item_detail.barang_data.nama_barang if item_detail.barang_data else '',
                        item_detail.jumlah,
                        item_detail.harga_satuan or 0.0, 
                        item_detail.subtotal or 0.0 
                    ]
                    cw.writerow(base_row + item_row_data)
            else: 
                 cw.writerow(base_row + [''] * 5) 
        else:
            cw.writerow(base_row)
            
    output = si.getvalue()
    si.close()
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename_prefix = "laporan_penjualan_detail_" if detail_item else "laporan_penjualan_"
    filename = f"{filename_prefix}{timestamp}.csv"

    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-disposition": f"attachment; filename={filename}"}
    )

# === XLSX Download ===

@sales_blueprint.route('/laporan_penjualan/download_xlsx', methods=['GET'], strict_slashes=False)
@login_required
def download_laporan_xlsx():
    """Generate interactive Excel (XLSX) sales report with filters and daily chart."""
    tanggal_mulai_str = request.args.get('tanggal_mulai')
    tanggal_akhir_str = request.args.get('tanggal_akhir')

    # Base query
    query = Penjualan.query.options(
        db.joinedload(Penjualan.pelanggan_data),
        db.joinedload(Penjualan.detail_penjualan).joinedload(DetailPenjualan.barang_data)
    ).order_by(Penjualan.tanggal.asc())

    if tanggal_mulai_str:
        try:
            tanggal_mulai = datetime.strptime(tanggal_mulai_str, '%Y-%m-%d').date()
            query = query.filter(Penjualan.tanggal >= datetime.combine(tanggal_mulai, datetime.min.time()))
        except ValueError:
            pass

    if tanggal_akhir_str:
        try:
            tanggal_akhir = datetime.strptime(tanggal_akhir_str, '%Y-%m-%d').date()
            query = query.filter(Penjualan.tanggal <= datetime.combine(tanggal_akhir, datetime.max.time()))
        except ValueError:
            pass

    transaksi_list = query.all()

    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Penjualan"

    headers = [
        'No Transaksi', 'Tanggal', 'Waktu', 'Pelanggan', 'Total Harga',
        'Metode Pembayaran', 'Jumlah Bayar', 'Kembalian', 'Status'
    ]
    ws.append(headers)

    for trx in transaksi_list:
        ws.append([
            trx.no_transaksi,
            trx.tanggal.strftime('%Y-%m-%d'),
            trx.tanggal.strftime('%H:%M:%S'),
            trx.pelanggan_data.nama if trx.pelanggan_data else 'Umum',
            float(trx.total_harga or 0.0),
            trx.metode_pembayaran or '',
            float(trx.jumlah_bayar) if trx.jumlah_bayar is not None else '',
            float(trx.kembalian) if trx.kembalian is not None else '',
            trx.status
        ])

    # Add auto filter
    ws.auto_filter.ref = f"A1:I{ws.max_row}"

    # Adjust column widths a bit
    for col in range(1, 10):
        ws.column_dimensions[chr(64+col)].width = 15

    # Create daily summary sheet
    sum_ws = wb.create_sheet("Ringkasan Harian")
    sum_ws.append(["Tanggal", "Total Penjualan", "Total Keuntungan", "Jumlah Transaksi"])

    # Build daily stats via Python to include profit
    daily_stats = {}
    for trx in transaksi_list:
        tgl_key = trx.tanggal.strftime('%Y-%m-%d')
        stat = daily_stats.setdefault(tgl_key, {'total': 0.0, 'profit': 0.0, 'count': 0})
        stat['total'] += float(trx.total_harga or 0.0)
        stat['count'] += 1
        # hitung profit transaksi
        profit_trx = 0.0
        for det in trx.detail_penjualan:
            modal = (det.barang_data.harga_pokok or 0.0) * det.jumlah if det.barang_data else 0.0
            profit_trx += float(det.subtotal or 0.0) - modal
        stat['profit'] += profit_trx

    # append to sheet ordered by date
    for tgl_key in sorted(daily_stats.keys()):
        stat = daily_stats[tgl_key]
        sum_ws.append([tgl_key, stat['total'], stat['profit'], stat['count']])

    # Add auto filter
    sum_ws.auto_filter.ref = f"A1:D{sum_ws.max_row}"

    if sum_ws.max_row > 2:
        # Chart for Total Penjualan
        chart_sales = BarChart()
        chart_sales.title = "Total Penjualan per Hari"
        chart_sales.y_axis.title = "Total Penjualan (Rp)"
        chart_sales.x_axis.title = "Tanggal"

        data_sales = Reference(sum_ws, min_col=2, min_row=1, max_row=sum_ws.max_row)
        cats = Reference(sum_ws, min_col=1, min_row=2, max_row=sum_ws.max_row)
        chart_sales.add_data(data_sales, titles_from_data=True)
        chart_sales.set_categories(cats)
        chart_sales.height = 10
        chart_sales.width = 20
        sum_ws.add_chart(chart_sales, "F2")

        # Chart for Profit
        chart_profit = BarChart()
        chart_profit.title = "Total Keuntungan per Hari"
        chart_profit.y_axis.title = "Keuntungan (Rp)"
        chart_profit.x_axis.title = "Tanggal"

        data_profit = Reference(sum_ws, min_col=3, min_row=1, max_row=sum_ws.max_row)
        chart_profit.add_data(data_profit, titles_from_data=True)
        chart_profit.set_categories(cats)
        chart_profit.height = 10
        chart_profit.width = 20
        sum_ws.add_chart(chart_profit, "F20")

    # Save workbook to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"laporan_penjualan_{timestamp}.xlsx"

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": f"attachment; filename={filename}"}
    )
