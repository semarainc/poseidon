import os
import uuid
import io
import traceback

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app
from datetime import datetime
from sqlalchemy import func, or_, inspect as sqlalchemy_inspect

from openpyxl import Workbook
import threading
import time
from werkzeug.utils import secure_filename

# Import dari paket models (menggunakan models/models.py)
from models.models import db, Barang, DetailPenjualan, Kategori, Satuan, Supplier
from flask_login import login_required, current_user

products_blueprint = Blueprint('products', __name__, template_folder='templates', static_folder='static', url_prefix='/products')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

# In-memory store for bulk upload job progress
# Structure: { job_id: { 'total': int, 'processed': int, 'inserted': int, 'updated': int, 'failed': int, 'done': bool, 'error': str|None } }
bulk_apply_jobs = {}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# --- BULK UPLOAD BARANG DARI EXCEL ---
@products_blueprint.route('/bulk_upload', methods=['GET'])
@login_required
def bulk_upload_barang():
    return render_template('barang_bulk_upload.html')

@products_blueprint.route('/download_template', methods=['GET'])
@login_required
def download_template():
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Barang'
    ws.append(['kode_barang', 'nama_barang', 'kategori', 'stok', 'stok_minimal_grosir', 'harga_pokok', 'harga_jual', 'harga_grosir', 'satuan', 'supplier', 'tanggal_kadaluarsa'])
    ws.append(['SKP001', 'Sekop Kecil', 'Alat Taman', 25, 24, 25000, 35000, 30000, 'Pcs', 'CV Taman Subur', '2025-12-31'])
    wb.save(output)
    output.seek(0)
    from flask import send_file
    return send_file(output, download_name='template_barang.xlsx', as_attachment=True)

@products_blueprint.route('/preview_bulk_upload', methods=['POST'])
@login_required
def preview_bulk_upload():
    import openpyxl
    file = request.files.get('excelFile')
    if not file:
        return jsonify({'success': False, 'message': 'File tidak ditemukan'}), 400
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() for h in rows[0]]
        data = []
        for row in rows[1:]:
            item = dict(zip(headers, row))
            data.append(item)
        return jsonify({'success': True, 'data': data, 'headers': headers})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Gagal parsing file: {str(e)}'}), 400

@products_blueprint.route('/apply_bulk_upload', methods=['POST'])
@login_required
def apply_bulk_upload():
    import datetime
    items = request.json.get('items', [])
    if not items:
        return jsonify({'success': False, 'message': 'Data barang kosong'}), 400
    inserted, updated, failed = 0, 0, 0
    for item in items:
        # Kategori
        kategori = (item.get('kategori') or '').strip()
        if kategori:
            kat_obj = Kategori.query.filter(func.lower(Kategori.nama) == kategori.lower()).first()
            if not kat_obj:
                kat_obj = Kategori(nama=kategori)
                db.session.add(kat_obj)
                db.session.commit()
        # Satuan
        satuan = (item.get('satuan') or '').strip()
        if satuan:
            sat_obj = Satuan.query.filter(func.lower(Satuan.nama) == satuan.lower()).first()
            if not sat_obj:
                sat_obj = Satuan(nama=satuan)
                db.session.add(sat_obj)
                db.session.commit()
        # Supplier
        supplier_nama = (item.get('supplier') or '').strip()
        supplier_obj = None
        supplier_id = None
        if supplier_nama:
            supplier_obj = Supplier.query.filter(func.lower(Supplier.nama) == supplier_nama.lower()).first()
            if not supplier_obj:
                try:
                    supplier_obj = Supplier(nama=supplier_nama)
                    db.session.add(supplier_obj)
                    db.session.flush()  # flush agar dapat id tanpa commit dulu
                except Exception:
                    db.session.rollback()
                    traceback.print_exc()
                    print("Error Occured: ", e)
                    # Cegah race condition: cek ulang (ada kemungkinan supplier ditambah oleh proses lain)
                    supplier_obj = Supplier.query.filter(func.lower(Supplier.nama) == supplier_nama.lower()).first()
                    if not supplier_obj:
                        failed += 1
                        continue
            supplier_id = supplier_obj.id
        # Barang
        try:
            kode_barang = (item.get('kode_barang') or '').strip()
            if not kode_barang:
                failed += 1
                continue
            barang = Barang.query.filter_by(kode_barang=kode_barang).first()
            tanggal_kadaluarsa = item.get('tanggal_kadaluarsa')
            if tanggal_kadaluarsa and isinstance(tanggal_kadaluarsa, str):
                try:
                    tanggal_kadaluarsa = datetime.datetime.strptime(tanggal_kadaluarsa, '%Y-%m-%d').date()
                except Exception:
                    tanggal_kadaluarsa = None
            if not barang:
                barang = Barang(
                    kode_barang=kode_barang,
                    nama_barang=item.get('nama_barang'),
                    kategori=kategori,
                    stok=float(item.get('stok') or 0),
                    batas_minimal_grosir=float(item.get('stok_minimal_grosir') or 0),
                    harga_pokok=float(item.get('harga_pokok') or 0),
                    harga_jual=float(item.get('harga_jual') or 0),
                    harga_grosir=float(item.get('harga_grosir') or 0),
                    satuan=satuan,
                    tanggal_kadaluarsa=tanggal_kadaluarsa,
                    supplier_id=supplier_id
                )
                db.session.add(barang)
                inserted += 1
            else:
                # Update data
                barang.nama_barang = item.get('nama_barang')
                barang.kategori = kategori
                barang.stok = float(item.get('stok') or 0)
                barang.batas_minimal_grosir = float(item.get('stok_minimal_grosir') or 0)
                barang.harga_pokok = float(item.get('harga_pokok') or 0)
                barang.harga_jual = float(item.get('harga_jual') or 0)
                barang.harga_grosir = float(item.get('harga_grosir') or 0)
                barang.satuan = satuan
                barang.tanggal_kadaluarsa = tanggal_kadaluarsa
                barang.supplier_id = supplier_id
                updated += 1
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            failed += 1
            print("Error Occured", e)
            traceback.print_exc()
    return jsonify({'success': True, 'inserted': inserted, 'updated': updated, 'failed': failed})

# --- BULK UPLOAD (ASYNCHRONOUS WITH PROGRESS) ---
@products_blueprint.route('/start_bulk_apply', methods=['POST'])
@login_required
def start_bulk_apply():
    """Start background job to apply bulk items and return a job_id to poll progress.
    """
    data = request.get_json(silent=True) or {}
    items = data.get('items', [])
    if not items:
        return jsonify({'success': False, 'message': 'Data barang kosong'}), 400

    job_id = str(uuid.uuid4())
    bulk_apply_jobs[job_id] = {
        'total': len(items),
        'processed': 0,
        'inserted': 0,
        'updated': 0,
        'failed': 0,
        'done': False,
        'error': None,
    }

    def worker(app_context, job_id, items):
        with app_context:
            inserted = updated = failed = 0
            processed = 0
            for item in items:
                try:
                    # --- replicate core logic from apply_bulk_upload ---
                    import datetime as _dt
                    # Kategori
                    kategori = (item.get('kategori') or '').strip()
                    if kategori:
                        kat_obj = Kategori.query.filter(func.lower(Kategori.nama) == kategori.lower()).first()
                        if not kat_obj:
                            kat_obj = Kategori(nama=kategori)
                            db.session.add(kat_obj)
                            db.session.commit()
                    # Satuan
                    satuan = (item.get('satuan') or '').strip()
                    if satuan:
                        sat_obj = Satuan.query.filter(func.lower(Satuan.nama) == satuan.lower()).first()
                        if not sat_obj:
                            sat_obj = Satuan(nama=satuan)
                            db.session.add(sat_obj)
                            db.session.commit()
                    # Supplier
                    supplier_nama = (item.get('supplier') or '').strip()
                    supplier_obj = None
                    supplier_id = None
                    if supplier_nama:
                        supplier_obj = Supplier.query.filter(func.lower(Supplier.nama) == supplier_nama.lower()).first()
                        if not supplier_obj:
                            try:
                                supplier_obj = Supplier(nama=supplier_nama)
                                db.session.add(supplier_obj)
                                db.session.flush()
                            except Exception:
                                db.session.rollback()
                                traceback.print_exc()
                                # Try fetch again to avoid race condition
                                supplier_obj = Supplier.query.filter(func.lower(Supplier.nama) == supplier_nama.lower()).first()
                                if not supplier_obj:
                                    failed += 1
                                    processed += 1
                                    bulk_apply_jobs[job_id].update({'processed': processed, 'inserted': inserted, 'updated': updated, 'failed': failed})
                                    continue
                        supplier_id = supplier_obj.id

                    # Barang
                    kode_barang = (item.get('kode_barang') or '').strip()
                    if not kode_barang:
                        failed += 1
                        processed += 1
                        bulk_apply_jobs[job_id].update({'processed': processed, 'inserted': inserted, 'updated': updated, 'failed': failed})
                        continue

                    barang = Barang.query.filter_by(kode_barang=kode_barang).first()
                    tanggal_kadaluarsa = item.get('tanggal_kadaluarsa')
                    if tanggal_kadaluarsa and isinstance(tanggal_kadaluarsa, str):
                        try:
                            tanggal_kadaluarsa = _dt.datetime.strptime(tanggal_kadaluarsa, '%Y-%m-%d').date()
                        except Exception:
                            tanggal_kadaluarsa = None

                    if not barang:
                        barang = Barang(
                            kode_barang=kode_barang,
                            nama_barang=item.get('nama_barang'),
                            kategori=kategori,
                            stok=float(item.get('stok') or 0),
                            batas_minimal_grosir=float(item.get('stok_minimal_grosir') or 0),
                            harga_pokok=float(item.get('harga_pokok') or 0),
                            harga_jual=float(item.get('harga_jual') or 0),
                            harga_grosir=float(item.get('harga_grosir') or 0),
                            satuan=satuan,
                            tanggal_kadaluarsa=tanggal_kadaluarsa,
                            supplier_id=supplier_id
                        )
                        db.session.add(barang)
                        inserted += 1
                    else:
                        barang.nama_barang = item.get('nama_barang')
                        barang.kategori = kategori
                        barang.stok = float(item.get('stok') or 0)
                        barang.batas_minimal_grosir = float(item.get('stok_minimal_grosir') or 0)
                        barang.harga_pokok = float(item.get('harga_pokok') or 0)
                        barang.harga_jual = float(item.get('harga_jual') or 0)
                        barang.harga_grosir = float(item.get('harga_grosir') or 0)
                        barang.satuan = satuan
                        barang.tanggal_kadaluarsa = tanggal_kadaluarsa
                        barang.supplier_id = supplier_id
                        updated += 1

                    try:
                        db.session.commit()
                    except Exception as e:
                        db.session.rollback()
                        failed += 1
                        current_app.logger.error(f"Bulk apply commit error: {e}")

                    processed += 1
                    # Update job progress
                    bulk_apply_jobs[job_id].update({
                        'processed': processed,
                        'inserted': inserted,
                        'updated': updated,
                        'failed': failed,
                    })

                    # Small sleep to avoid starving event loop (optional)
                    time.sleep(0.001)
                except Exception as e:
                    # Catch any unexpected per-item error
                    db.session.rollback()
                    failed += 1
                    processed += 1
                    current_app.logger.error(f"Bulk apply item error: {e}")
                    bulk_apply_jobs[job_id].update({
                        'processed': processed,
                        'inserted': inserted,
                        'updated': updated,
                        'failed': failed,
                        'error': str(e),
                    })

            # Mark done
            bulk_apply_jobs[job_id]['done'] = True

    # Launch background thread
    t = threading.Thread(target=worker, args=(current_app.app_context(), job_id, items), daemon=True)
    t.start()

    return jsonify({'success': True, 'job_id': job_id})

@products_blueprint.route('/bulk_apply_status/<job_id>', methods=['GET'])
@login_required
def bulk_apply_status(job_id):
    job = bulk_apply_jobs.get(job_id)
    if not job:
        return jsonify({'success': False, 'message': 'Job tidak ditemukan'}), 404
    return jsonify({'success': True, **job})

# --- MANAJEMEN BARANG ---
@products_blueprint.route('/', strict_slashes=False)
@login_required
def barang():
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    allowed_sizes = [10, 25, 50, 100]
    if per_page not in allowed_sizes:
        per_page = 25

    query = Barang.query
    if search:
        query = query.filter(
            or_(Barang.nama_barang.contains(search), Barang.kode_barang.contains(search))
        )

    total_count = query.count()
    pagination = query.order_by(Barang.nama_barang).paginate(page=page, per_page=per_page, error_out=False)
    barang_list = pagination.items
    return render_template('barang.html', barang_list=barang_list, pagination=pagination, search=search, per_page=per_page, total_count=total_count)

@products_blueprint.route('/tambah', methods=['GET', 'POST'], strict_slashes=False)
@login_required
def tambah_barang():
    supplier_list = Supplier.query.order_by(Supplier.nama).all()
    kategori_list = [k.nama for k in Kategori.query.order_by(Kategori.nama).all()]
    satuan_list   = [s.nama for s in Satuan.query.order_by(Satuan.nama).all()]
    if request.method == 'POST':
        kode_barang = request.form['kode_barang']
        nama_barang = request.form['nama_barang']
        kategori = request.form['kategori']
        satuan = request.form['satuan']
        supplier_id = request.form.get('supplier_id') or None
        stok = int(request.form.get('stok', 0))
        harga_pokok = float(request.form.get('harga_pokok', 0.0))
        harga_jual = float(request.form.get('harga_jual', 0.0))
        harga_grosir = float(request.form.get('harga_grosir', 0.0))
        batas_minimal_grosir = int(request.form.get('batas_minimal_grosir', 0))
        
        tanggal_kadaluarsa_str = request.form.get('tanggal_kadaluarsa')
        tanggal_kadaluarsa_obj = None
        if tanggal_kadaluarsa_str:
            try:
                tanggal_kadaluarsa_obj = datetime.strptime(tanggal_kadaluarsa_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Format tanggal kadaluarsa tidak valid. Gunakan YYYY-MM-DD.', 'error')
                return render_template('barang_form.html', action='Tambah', barang=request.form, supplier_list=supplier_list, kategori_list=kategori_list, satuan_list=satuan_list)

        gambar_filename = None
        if 'gambar' in request.files:
            file = request.files['gambar']
            if file and file.filename != '' and allowed_file(file.filename):
                ext = file.filename.rsplit('.', 1)[1].lower()
                unique_filename = str(uuid.uuid4()) + '.' + ext
                filename = secure_filename(unique_filename)
                file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
                gambar_filename = filename
            elif file and file.filename != '':
                flash('Format file gambar tidak diizinkan! (Hanya PNG, JPG, JPEG, GIF)', 'error')
                return render_template('barang_form.html', action='Tambah', barang=request.form, supplier_list=supplier_list, kategori_list=kategori_list, satuan_list=satuan_list)

        existing = Barang.query.filter_by(kode_barang=kode_barang).first()
        if existing:
            flash('Kode barang sudah ada!', 'error')
            return render_template('barang_form.html', action='Tambah', barang=request.form, supplier_list=supplier_list, kategori_list=kategori_list, satuan_list=satuan_list)
        
        new_barang = Barang(
            kode_barang=kode_barang,
            nama_barang=nama_barang,
            satuan=satuan,
            kategori=kategori,
            stok=stok,
            harga_jual=harga_jual,
            harga_pokok=harga_pokok,
            harga_grosir=harga_grosir,
            batas_minimal_grosir=batas_minimal_grosir,
            supplier_id=supplier_id,
            gambar=gambar_filename,
            tanggal_kadaluarsa=tanggal_kadaluarsa_obj
        )
        try:
            db.session.add(new_barang)
            db.session.commit()
            flash('Barang berhasil ditambahkan!', 'success')
            return redirect(url_for('products.barang'))
        except Exception as e:
            db.session.rollback()
            if gambar_filename and os.path.exists(os.path.join(current_app.config['UPLOAD_FOLDER'], gambar_filename)):
                 os.remove(os.path.join(current_app.config['UPLOAD_FOLDER'], gambar_filename))
            flash(f'Gagal menambahkan barang: {str(e)}', 'error')
            current_app.logger.error(f"Error tambah barang: {e}")
            return render_template('barang_form.html', action='Tambah', barang=request.form, supplier_list=supplier_list, kategori_list=kategori_list, satuan_list=satuan_list)
    
    return render_template('barang_form.html', action='Tambah', barang=None, supplier_list=supplier_list, kategori_list=kategori_list, satuan_list=satuan_list)

@products_blueprint.route('/edit/<int:id>', methods=['GET', 'POST'], strict_slashes=False)
@login_required
def edit_barang(id):
    supplier_list = Supplier.query.order_by(Supplier.nama).all()
    kategori_list = [k.nama for k in Kategori.query.order_by(Kategori.nama).all()]
    satuan_list   = [s.nama for s in Satuan.query.order_by(Satuan.nama).all()]
    barang_item = Barang.query.get_or_404(id)
    if request.method == 'POST':
        new_kode_barang = request.form['kode_barang']
        if new_kode_barang != barang_item.kode_barang:
            existing = Barang.query.filter(Barang.id != id, Barang.kode_barang == new_kode_barang).first()
            if existing:
                flash('Kode barang sudah digunakan oleh barang lain!', 'error')
                return render_template('barang_form.html', action='Edit', barang=barang_item, supplier_list=supplier_list, kategori_list=kategori_list, satuan_list=satuan_list)

        tanggal_kadaluarsa_str = request.form.get('tanggal_kadaluarsa')
        tanggal_kadaluarsa_obj = None 
        if tanggal_kadaluarsa_str: 
            try:
                tanggal_kadaluarsa_obj = datetime.strptime(tanggal_kadaluarsa_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Format tanggal kadaluarsa tidak valid.', 'error')
                return render_template('barang_form.html', action='Edit', barang=barang_item, supplier_list=supplier_list, kategori_list=kategori_list, satuan_list=satuan_list)
        
        gambar_filename = barang_item.gambar
        if 'gambar' in request.files:
            file = request.files['gambar']
            if file and file.filename != '' and allowed_file(file.filename):
                if barang_item.gambar and os.path.exists(os.path.join(current_app.config['UPLOAD_FOLDER'], barang_item.gambar)):
                    try:
                        os.remove(os.path.join(current_app.config['UPLOAD_FOLDER'], barang_item.gambar))
                    except OSError as e:
                        current_app.logger.error(f"Error deleting old image: {e}")
                ext = file.filename.rsplit('.', 1)[1].lower()
                unique_filename = str(uuid.uuid4()) + '.' + ext
                filename = secure_filename(unique_filename)
                file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
                gambar_filename = filename
            elif file and file.filename != '': 
                flash('Format file gambar tidak diizinkan! (Hanya PNG, JPG, JPEG, GIF)', 'error')
                return render_template('barang_form.html', action='Edit', barang=barang_item, supplier_list=supplier_list, kategori_list=kategori_list, satuan_list=satuan_list)

        supplier_id = request.form.get('supplier_id') or None
        barang_item.kode_barang = new_kode_barang
        barang_item.nama_barang = request.form['nama_barang']
        barang_item.satuan = request.form['satuan']
        barang_item.kategori = request.form['kategori']
        barang_item.stok = float(request.form.get('stok', barang_item.stok))
        barang_item.harga_jual = float(request.form.get('harga_jual', barang_item.harga_jual))
        barang_item.harga_pokok = float(request.form.get('harga_pokok', barang_item.harga_pokok))
        barang_item.harga_grosir = float(request.form.get('harga_grosir', barang_item.harga_grosir))
        barang_item.batas_minimal_grosir = int(request.form.get('batas_minimal_grosir', barang_item.batas_minimal_grosir))
        barang_item.supplier_id = supplier_id
        barang_item.gambar = gambar_filename
        barang_item.tanggal_kadaluarsa = tanggal_kadaluarsa_obj 
        
        try:
            db.session.commit()
            flash('Barang berhasil diperbarui!', 'success')
            return redirect(url_for('products.barang'))
        except Exception as e:
            db.session.rollback()
            traceback.print_exc()
            flash(f'Gagal memperbarui barang: {str(e)}', 'error')
            current_app.logger.error(f"Error edit barang: {e}")
            return render_template('barang_form.html', action='Edit', barang=barang_item, supplier_list=supplier_list, kategori_list=kategori_list, satuan_list=satuan_list)
    return render_template('barang_form.html', action='Edit', barang=barang_item, supplier_list=supplier_list, kategori_list=kategori_list, satuan_list=satuan_list)

@products_blueprint.route('/hapus/<int:id>', strict_slashes=False)
@login_required
def hapus_barang(id):
    barang_item = Barang.query.get_or_404(id)
    try:
        if DetailPenjualan.query.filter_by(barang_id=id).first():
            flash('Barang tidak bisa dihapus karena sudah ada dalam transaksi penjualan.', 'error')
            return redirect(url_for('products.barang'))
        gambar_path = None
        if barang_item.gambar:
             gambar_path = os.path.join(current_app.config['UPLOAD_FOLDER'], barang_item.gambar)
        db.session.delete(barang_item)
        db.session.commit()
        if gambar_path and os.path.exists(gambar_path):
            try:
                os.remove(gambar_path)
            except OSError as e:
                current_app.logger.error(f"Error deleting image file: {e}")
        flash('Barang berhasil dihapus!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menghapus barang: {str(e)}', 'error')
        current_app.logger.error(f"Error hapus barang: {e}")
    return redirect(url_for('products.barang'))

# --- API untuk Kategori dan Satuan ---
@products_blueprint.route('/api/kategori', methods=['GET'], strict_slashes=False)
@login_required
def get_all_kategori():
    kategori_list = Kategori.query.order_by(Kategori.nama).all()
    return jsonify({'kategori': [kategori.to_dict() for kategori in kategori_list]})

@products_blueprint.route('/api/kategori', methods=['POST'], strict_slashes=False)
@login_required
def add_kategori():
    data = request.get_json()
    if not data or 'nama' not in data:
        return jsonify({'success': False, 'message': 'Nama kategori harus diisi'}), 400
        
    nama = data['nama'].strip()
    if not nama:
        return jsonify({'success': False, 'message': 'Nama kategori tidak boleh kosong'}), 400
    
    existing = Kategori.query.filter(func.lower(Kategori.nama) == func.lower(nama)).first()
    if existing:
        return jsonify({'success': False, 'message': 'Kategori dengan nama ini sudah ada'}), 400
    
    try:
        new_kategori = Kategori(nama=nama)
        db.session.add(new_kategori)
        db.session.commit()
        return jsonify({
            'success': True, 
            'message': 'Kategori berhasil ditambahkan', 
            'kategori': new_kategori.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error adding kategori: {e}")
        return jsonify({'success': False, 'message': f'Gagal menambahkan kategori: {str(e)}'}), 500

@products_blueprint.route('/api/kategori/<int:id>', methods=['PUT'], strict_slashes=False)
@login_required
def update_kategori(id):
    kategori = Kategori.query.get_or_404(id)
    data = request.get_json()
    
    if not data or 'nama' not in data:
        return jsonify({'success': False, 'message': 'Nama kategori harus diisi'}), 400
        
    nama = data['nama'].strip()
    if not nama:
        return jsonify({'success': False, 'message': 'Nama kategori tidak boleh kosong'}), 400
    
    existing = Kategori.query.filter(Kategori.id != id, func.lower(Kategori.nama) == func.lower(nama)).first()
    if existing:
        return jsonify({'success': False, 'message': 'Kategori dengan nama ini sudah ada'}), 400
    
    try:
        kategori.nama = nama
        db.session.commit()
        return jsonify({
            'success': True, 
            'message': 'Kategori berhasil diperbarui', 
            'kategori': kategori.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating kategori: {e}")
        return jsonify({'success': False, 'message': f'Gagal memperbarui kategori: {str(e)}'}), 500

@products_blueprint.route('/api/kategori/<int:id>', methods=['DELETE'], strict_slashes=False)
@login_required
def delete_kategori(id):
    kategori = Kategori.query.get_or_404(id)
       
    # Check if kategori is being used in any products
    used_in_products = Barang.query.filter_by(kategori=kategori.nama).first()
    if used_in_products:
        return jsonify({
            'success': False, 
            'message': f'Kategori sedang digunakan oleh barang: {used_in_products.nama_barang}'
        }), 400
    
    try:
        db.session.delete(kategori)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Kategori berhasil dihapus'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting kategori: {e}")
        return jsonify({'success': False, 'message': f'Gagal menghapus kategori: {str(e)}'}), 500

# Satuan (Unit) API
@products_blueprint.route('/api/satuan', methods=['GET'], strict_slashes=False)
@login_required
def get_all_satuan():
    satuan_list = Satuan.query.order_by(Satuan.nama).all()
    return jsonify({'satuan': [satuan.to_dict() for satuan in satuan_list]})

@products_blueprint.route('/api/satuan', methods=['POST'])
@login_required
def add_satuan():
    data = request.get_json()
    if not data or 'nama' not in data:
        return jsonify({'success': False, 'message': 'Nama satuan harus diisi'}), 400
        
    nama = data['nama'].strip()
    deskripsi = data.get('deskripsi', '').strip()
    
    if not nama:
        return jsonify({'success': False, 'message': 'Nama satuan tidak boleh kosong'}), 400
    
    existing = Satuan.query.filter(func.lower(Satuan.nama) == func.lower(nama)).first()
    if existing:
        return jsonify({'success': False, 'message': 'Satuan dengan nama ini sudah ada'}), 400
    
    try:
        new_satuan = Satuan(nama=nama, deskripsi=deskripsi)
        db.session.add(new_satuan)
        db.session.commit()
        return jsonify({
            'success': True, 
            'message': 'Satuan berhasil ditambahkan', 
            'satuan': new_satuan.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error adding satuan: {e}")
        return jsonify({'success': False, 'message': f'Gagal menambahkan satuan: {str(e)}'}), 500

@products_blueprint.route('/api/satuan/<int:id>', methods=['PUT'], strict_slashes=False)
@login_required
def update_satuan(id):
    satuan = Satuan.query.get_or_404(id)
    data = request.get_json()
    
    if not data or 'nama' not in data:
        return jsonify({'success': False, 'message': 'Nama satuan harus diisi'}), 400
        
    nama = data['nama'].strip()
    deskripsi = data.get('deskripsi', satuan.deskripsi or '').strip()
    
    if not nama:
        return jsonify({'success': False, 'message': 'Nama satuan tidak boleh kosong'}), 400
    
    existing = Satuan.query.filter(Satuan.id != id, func.lower(Satuan.nama) == func.lower(nama)).first()
    if existing:
        return jsonify({'success': False, 'message': 'Satuan dengan nama ini sudah ada'}), 400
    
    try:
        satuan.nama = nama
        satuan.deskripsi = deskripsi
        db.session.commit()
        return jsonify({
            'success': True, 
            'message': 'Satuan berhasil diperbarui', 
            'satuan': satuan.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating satuan: {e}")
        return jsonify({'success': False, 'message': f'Gagal memperbarui satuan: {str(e)}'}), 500

@products_blueprint.route('/api/satuan/<int:id>', methods=['DELETE'], strict_slashes=False)
@login_required
def delete_satuan(id):
    satuan = Satuan.query.get_or_404(id)
    
    # Check if satuan is being used in any products
    used_in_products = Barang.query.filter_by(satuan=satuan.nama).first()
    if used_in_products:
        return jsonify({
            'success': False, 
            'message': f'Satuan sedang digunakan oleh barang: {used_in_products.nama_barang}'
        }), 400
    
    try:
        db.session.delete(satuan)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Satuan berhasil dihapus'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting satuan: {e}")
        return jsonify({'success': False, 'message': f'Gagal menghapus satuan: {str(e)}'}), 500
